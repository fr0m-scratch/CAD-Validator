from utility.utility import *
from cad_entity import CADEntity
from cad_handler import CADHandler
import pandas as pd
import openpyxl
from utility import*
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, colors
from rich.progress import Progress

#CAD Info Extration
acad = CADHandler("entity_list.pkl")
acad.set_params()
at,b,c = acad.identify_accessories()
illustration = acad.retrive_designation()

#Comparing Data Unification
control, ref = read_sheet('control.xlsx')
extracted = acad.accessories_to_df(at, b,c)
for c, e in zip(control, extracted):
    unify(c)
    unify(e)
sensors, specials, actuators = extracted
sensors_control, specials_control, actuators_control = control

def mark_diff(filepath):
    wb = openpyxl.load_workbook(filepath)
    sensors_writer = wb['SENSOR_IO']
    specials_writer = wb['SPECIAL_IO']
    actuators_writer = wb['ACTUATOR_IO']
    for row, col in read_diff(sensors, sensors_control,ref, extension=False)[0]:
        sensors_writer.cell(row, col).font = Font(color='FF0000')
    for row, col in read_diff(specials, specials_control,ref,extension=False)[0]:
        specials_writer.cell(row, col).font = Font(color='FF0000')
    for row, col in read_diff(actuators, actuators_control,ref)[0]:
        actuators_writer.cell(row, col).font = Font(color='FF0000')
    
    wb.save(filepath)
mark_diff('control.xlsx')

def write_diff(filepath):
    return