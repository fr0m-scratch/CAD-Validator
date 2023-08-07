import re
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Color, Alignment, Border, Side, PatternFill
from rich.progress import Progress
import os
import sys
import ezdxf 
import difflib
import math
def get_match_ratio(a, b, ignore_list):
    return difflib.SequenceMatcher(lambda x: x in ignore_list, a, b).quick_ratio()
def find_graph_id(coordinates, rectangles):
    #coordinates is a tuple of x and y
    #rectangles 是一个列表，每个元素是一个元组，元组的第一个元素是一个四元组，分别是左下角和右上角的坐标，第二个元素是图形的id
    for i in range(len(rectangles)):
        x1, x2, y1, y2 = rectangles[i][0]
        graph_id = rectangles[i][1].tag
        x, y = coordinates[0], coordinates[1]
        if x1 <= x <= x2 and y1 <= y <= y2:
            return graph_id, rectangles[i]
    return None, None

def parameter_retriving(entity):
    """entity is a dxf entity, parameters is a list of parameters to be retrived"""
    #entity是一个dxf实体，parameters是一个列表，每个元素是一个参数名，返回一个字典，key是参数名，value是参数值
    relevant_attributes = ['ObjectName', 'InsertionPoint', 'coordinates', 'Area', 'Length', 'TextString', 'Handle', 'height', 'Layer', 'TagString', 'color']
    relevant_member_function = ['Coordinate', 'GetBoundingBox']
    res_dict = {}
    for function in relevant_member_function:
        try: 
            res = getattr(entity, function)()
            res_dict[function] = res
        except:
            res_dict[function] = None
    for attribute in relevant_attributes:
        try:
            attr = getattr(entity, attribute)
            res_dict[attribute] = attr
        except:
            res_dict[attribute] = None
    if res_dict['coordinates'] is not None:
        coordinates = []
        for i in range(len(res_dict['coordinates'])//2):
            coordinates.append(entity.Coordinate(i))
        res_dict['coordinates'] = coordinates
    return res_dict

def remove_entity_by_string(entity_list, string):
    #去除entity_list中text中包含string的entity
    for entity in entity_list:
        if re.search(string, entity.text) is not None:
            entity_list.remove(entity)
    return entity_list

def enum_to_dict(enum):
    #把enum转换成字典，key是序号，value是值
    return {v:k for k,v in enum}

def read_sheet(filepath):
    #读取excel表格，返回一个元组，元组的每个元素是一个dataframe，分别是SENSOR_IO, SPECIAL_IO, ACTUATOR_IO
    with Progress() as progress:
        read = progress.add_task("[cyan2]Read[cyan2]file", total=14)
        sensors_control = pd.read_excel(filepath, sheet_name='SENSOR_IO')
        progress.update(read, advance=4)
        specials_control = pd.read_excel(filepath, sheet_name='SPECIAL_IO')
        progress.update(read, advance=5)
        actuators_control = pd.read_excel(filepath, sheet_name='ACTUATOR_IO')
        progress.update(read, advance=1)
        for _ in (sensors_control, specials_control, actuators_control):
            _.rename(mapper= lambda x: ''.join(''.join(''.join(x.split('_x000D_\n')).split('\n')).split('-')), axis=1, inplace=True)
            progress.update(read, advance=1)
        ref = enum_to_dict(enumerate(sensors_control.columns.tolist(),1))
        progress.update(read, advance=1)
    return (sensors_control, specials_control, actuators_control), ref

def unify(df):
    #统一dataframe中的数据顺序以进行比对
    df.sort_values(by=['信号位号idcode', '图号diagram number','扩展码extensioncode'], inplace=True, ignore_index=True)
    return df
def read_diff(control_dataframe,extract_dict, ref, extension, designation):
        diff = []
        diff_content = {}
        for control in control_dataframe.iterrows():
            control = control[1]
            if not extension:
                try:
                    key = (control['信号位号idcode'][1:], control['图号diagram number'])
                except:
                    continue
            else:
                try:
                    key = (control['信号位号idcode'][1:], control['扩展码extensioncode'], control.loc['图号diagram number'])
                except:
                    continue
            if key in extract_dict:
                check_diff(control, extract_dict[key], ref,  diff, diff_content, extension, designation)
                if extract_dict[key].relevance and extract_dict[key].relevance.exception:
                    diff.append((ref['信号位号idcode'], control['refindex']))
                    diff_content[('信号位号idcode', control['refindex'])] = extract_dict[key].relevance.exception.mismatch[extract_dict[key]['信号位号idcode']]
            else:
                diff.append(('all', control['refindex']))
        return diff, diff_content
    
def check_diff(control, extract, ref,  diff, diff_content,extension=False, designation=True,):
    labels = ['扩展码extensioncode','信号位号idcode', '图号diagram number', '版本rev.', '信号说明designation', "安全分级/分组Safetyclass/division"]
    check = labels.copy()
    if not designation:
        check.remove('信号说明designation')
    if not extension:
        check.remove('扩展码extensioncode')
    for col in check:
        if col == '信号位号idcode':
            diff_check = control[col][1:] != extract[col][1:]
        else:
            diff_check = control[col] != extract[col]
        
        if diff_check:
            if col == '信号说明designation':
                if abbreviation_match(control,extract):
                    continue
            col_name, col, row, content = col, ref[col], control['refindex'], extract[col]
            diff.append((col, row))
            diff_content[(col_name, row)] = content
    return None

def output_diff(filepath, ref, comparing_data):
        #把比对结果写入excel表格
        sensors, specials, actuators, sensors_control, specials_control, actuators_control = comparing_data
        wb = openpyxl.load_workbook(filepath)
        sensors_writer = wb['SENSOR_IO']
        specials_writer = wb['SPECIAL_IO']
        actuators_writer = wb['ACTUATOR_IO']
        col_dict = {
            "信号位号idcode": (1, "CAD信号位号"),
            "扩展码extensioncode": (2, "CAD扩展码"),
            "信号说明designation": (3, "CAD信号说明"),
            "安全分级/分组Safetyclass/division": (4, "CAD安全分级/分组"),
            "图号diagram number": (5, "CAD图号"),
            "版本rev.": (6, "CAD版本"),
        }
        
        def write_diff(writer, extracted, control, extension = True, designation = True ):
            pos, content = read_diff(control, extracted, ref, extension, designation)
            insert_col = writer.max_column + 3
            writer.insert_cols(insert_col, 1)
            for col, row in pos:
                if col == 'all':
                    for column in range(1, writer.max_column+1):
                        writer.cell(row, column).font = Font(color='FF0000')
                    writer.cell(row, insert_col+1).value = 'CAD图纸中不存在该信号位号'
                    writer.cell(row, insert_col+1).font = Font(color='FF0000')
                else:
                    writer.cell(row, col).font = Font(color='FF0000')
            for col_name, row in content.keys():
                writer.cell(row, insert_col+col_dict[col_name][0]).value = str(content[(col_name, row)])
            for col in col_dict.values():
                writer.cell(1, insert_col+col[0]).value = col[1]
                writer.cell(1, insert_col+col[0]).font = Font(bold=True, size=10)
                writer.cell(1, insert_col+col[0]).alignment = Alignment(horizontal='center', vertical='center')
            for i in range(insert_col, insert_col+5):
                writer.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 15
            return
        write_diff(sensors_writer, sensors, sensors_control, extension = False)
        write_diff(specials_writer, specials, specials_control, extension = False)
        write_diff(actuators_writer, actuators, actuators_control)
        wb.save(filepath)

def get_resource_path(relative_path):
    """解析编译后的数据文件路径"""
    if hasattr(sys, '_MEIPASS'):
        return os.path.join(sys._MEIPASS, relative_path)
    return os.path.join(os.path.abspath("."), relative_path)

def contain(words,entities):
        #根据关键词查找实体
        res = []
        for entity in entities:
            if entity.text is not None:
                text = re.search(words, entity.text)
                if text is not None:
                    res.append(entity)
            if entity.tag is not None:
                tag = re.search(words, entity.tag)
                if tag is not None:
                    res.append(entity)
        return res
    
def not_contain(words, entities):
        #根据关键词查找实体
        res = []
        for entity in entities:
            if entity.text is not None:
                text = re.search(words, entity.text)
                if text is None:
                    res.append(entity)
            if entity.tag is not None:
                tag = re.search(words, entity.tag)
                if tag is None:
                    res.append(entity)
        return set(res)
    
def unformat_mtext(s, exclude_list=('n', 'S')):
    """Returns string with removed format information
    :param s: string with multitext
    :param exclude_list: don't touch tags from this list. Default ('P', 'S') for
                         newline and fractions
    >>> text = ur'{\\fGOST type A|b0|i0|c204|p34;TEST\\fGOST type A|b0|i0|c0|p34;123}'
    >>> unformat_mtext(text)
    u'TEST123'
    """
    s = re.sub(r'\{?\\[^%s][^;]+;' % ''.join(exclude_list), '', s)
    s = re.sub(r'\}', '', s)
    return s



def mtext_to_string(s):
    """
    Returns string with removed format innformation as :func:`unformat_mtext` and
    `\\P` (paragraphs) replaced with newlines
    >>> text = ur'{\\fGOST type A|b0|i0|c204|p34;TEST\\fGOST type A|b0|i0|c0|p34;123}\\Ptest321'
    >>> mtext_to_string(text)
    u'TEST123\\ntest321'
    """
    return unformat_mtext(s)


def abbrev_match(abr, ori, threshold = 1000):
    i,j, count = 0, 0, 0
    while i < len(abr) and j < len(ori):
        if abr[i] == ori[j]:
            i += 1
            j += 1
        else:
            j += 1
            count += 1
    count += len(ori) - j
    if count >= threshold:
        return False
    if i == len(abr):
        return True
    elif j == len(ori):
        return False

def abbreviation_match(control, extract):
    
    if control['信号说明designation'].strip() == '':
        return False
    # if control['信号位号idcode'] == '3TFR301VL':
    #     print(control['信号说明designation'], extract['信号说明designation'])
    control_designation = control['信号说明designation'].split('\n')
    if control['信号说明designation'] == '' and extract['信号说明designation'] != '无描述':
        return False
    if get_match_ratio(control['信号说明designation'], extract['信号说明designation'], ['\n', ' ']) > 0.8:
        return True
    if len(control_designation) == 1:
        control_designation = [_ for _ in control['信号说明designation'].split('    ') if _ != '']
    c_chinese = ''.join([x for x in control_designation if re.search('[\u4e00-\u9fff]', x)])
    c_english = ''.join([x for x in control_designation if not re.search('[\u4e00-\u9fff]', x)])
    # if control['信号位号idcode'] == '3TFR301VL':
    #         print(control_designation, c_chinese, c_english)
    for designation in extract.designation_list:
        designation.text = designation.text.replace('\\P', '')
    chinese = [x for x in extract.designation_list if (re.search('[\u4e00-\u9fff]', x.text) and (not re.search('控制柜', x.text)))]
    english = [x for x in extract.designation_list if not re.search('[\u4e00-\u9fff]', x.text)]
    # if control['信号位号idcode'] == '3TFR301VL':
    #         print(control_designation, chinese, english)
    english_ratio = [(get_match_ratio(c_english, e_english.text, ['\n', ' ']), e_english) for e_english in english]
    chinese_ratio = [(get_match_ratio(c_chinese, e_chinese.text, ['\n', ' ']), e_chinese) for e_chinese in chinese]
    if chinese_ratio:
        # if control['信号位号idcode'] == '3TFP200SY':
        #     print(chinese_ratio)
        chinese = max(chinese_ratio, key=lambda x: x[0])
    else:
        chinese = (0, extract)
    if english_ratio:
        english = max(english_ratio, key=lambda x: x[0])
    else:
        english = (0, extract)
    if chinese[0] > 0.8 or english[0] > 0.8:
        return True
    e_chinese, e_english = chinese[1], english[1]
    
        
    chinese_check = (c_chinese and e_chinese.text) and ((abbrev_match(c_chinese, e_chinese.text) or abbrev_match(e_chinese.text, c_chinese)))
    english_check = (c_english and e_english.text) and ((abbrev_match(c_english, e_english.text) or abbrev_match(e_english.text, c_english)))
    if (chinese_check or english_check) and extract['信号说明designation'] != '无描述':
        return True
    return False
    
    
    
    
def get_arc_length_area(e):
    end_a = e.dxf.end_angle
    start_a=e.dxf.start_angle
    r=e.dxf.radius
    #弧长l = A * np.pi * R / 180
    a=end_a+(360 if start_a>end_a else 0)-start_a
    l=a*r*math.pi/180
    #圆弧面积
    area=l*r/2
    #CAD圆弧面积=圆弧面积 加或减 三角面积
    arc_delta=ezdxf.math.area([e.start_point,e.end_point,e.dxf.center])
    cad_area=area - (arc_delta if end_a-start_a<180 else -1*arc_delta)
    #圆弧中点坐标
    x= e.dxf.center[0]+r*math.cos(math.radians(a/2+start_a));
    y= e.dxf.center[1]+r*math.sin(math.radians(a/2+start_a));
    arc_mid=x,y
    return l,area,cad_area,arc_mid
